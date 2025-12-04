# # """
# # Generates XLSX output files for comparison results.
# # Creates separate files for table and text differences.
# # """

# # from pathlib import Path
# # from typing import List, Dict
# # import openpyxl
# # from openpyxl.styles import Font, PatternFill, Alignment
# # from datetime import datetime


# # class OutputGenerator:
# #     """Generates formatted XLSX output files."""
    
# #     def __init__(self):
# #         """Initialize with style definitions."""
# #         self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
# #         self.header_font = Font(bold=True, color="FFFFFF")
# #         self.alternating_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
# #     def generate_table_output(self, table_differences: List[Dict], output_file: str):
# #         """
# #         Generate XLSX file with table differences.
        
# #         Args:
# #             table_differences: List of table difference records
# #             output_file: Path to output XLSX file
# #         """
        
# #         workbook = openpyxl.Workbook()
# #         sheet = workbook.active
# #         sheet.title = "Table Differences"
        
# #         # Write headers
# #         headers = ["Table Index", "Row Index", "File From", "Cell Values", "Status"]
# #         for col_num, header in enumerate(headers, 1):
# #             cell = sheet.cell(row=1, column=col_num)
# #             cell.value = header
# #             cell.fill = self.header_fill
# #             cell.font = self.header_font
# #             cell.alignment = Alignment(horizontal="center", vertical="center")
        
# #         # Write data rows
# #         row_num = 2
# #         for table_diff in table_differences:
# #             for diff in table_diff['differences']:
# #                 sheet.cell(row=row_num, column=1).value = table_diff['table_index']
# #                 sheet.cell(row=row_num, column=2).value = diff['row_index']
# #                 sheet.cell(row=row_num, column=3).value = diff['file_from']
# #                 sheet.cell(row=row_num, column=4).value = " | ".join(diff['values'])
# #                 sheet.cell(row=row_num, column=5).value = diff['status']
                
# #                 # Apply alternating colors for readability
# #                 if row_num % 2 == 0:
# #                     for col in range(1, len(headers) + 1):
# #                         sheet.cell(row=row_num, column=col).fill = self.alternating_fill
                
# #                 row_num += 1
        
# #         # Auto-adjust column widths
# #         self._auto_adjust_columns(sheet)
        
# #         workbook.save(output_file)
# #         print(f"Table output saved to {output_file}")
    
# #     def generate_text_output(self, text_differences: List[Dict], output_file: str):
# #         """
# #         Generate XLSX file with text differences.
        
# #         Args:
# #             text_differences: List of text difference records
# #             output_file: Path to output XLSX file
# #         """
        
# #         workbook = openpyxl.Workbook()
# #         sheet = workbook.active
# #         sheet.title = "Text Differences"
        
# #         # Write headers
# #         headers = ["Paragraph Index", "File From", "Text Content", "Status"]
# #         for col_num, header in enumerate(headers, 1):
# #             cell = sheet.cell(row=1, column=col_num)
# #             cell.value = header
# #             cell.fill = self.header_fill
# #             cell.font = self.header_font
# #             cell.alignment = Alignment(horizontal="center", vertical="center")
        
# #         # Write data rows
# #         row_num = 2
# #         for diff in text_differences:
# #             sheet.cell(row=row_num, column=1).value = diff['paragraph_index']
# #             sheet.cell(row=row_num, column=2).value = diff['file_from']
# #             sheet.cell(row=row_num, column=3).value = diff['text']
# #             sheet.cell(row=row_num, column=4).value = diff['status']
            
# #             # Apply alternating colors for readability
# #             if row_num % 2 == 0:
# #                 for col in range(1, len(headers) + 1):
# #                     sheet.cell(row=row_num, column=col).fill = self.alternating_fill
            
# #             # Wrap text in content column
# #             sheet.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            
# #             row_num += 1
        
# #         # Auto-adjust column widths
# #         self._auto_adjust_columns(sheet)
        
# #         workbook.save(output_file)
# #         print(f"Text output saved to {output_file}")
    
# #     def _auto_adjust_columns(self, sheet):
# #         """
# #         Auto-adjust column widths based on content.
        
# #         Args:
# #             sheet: Openpyxl worksheet object
# #         """
        
# #         for column in sheet.columns:
# #             max_length = 0
# #             column_letter = column[0].column_letter
            
# #             for cell in column:
# #                 try:
# #                     if cell.value:
# #                         cell_length = len(str(cell.value))
# #                         if cell_length > max_length:
# #                             max_length = cell_length
# #                 except:
# #                     pass
            
# #             # Set column width with some padding
# #             adjusted_width = min(max_length + 2, 50)
# #             sheet.column_dimensions[column_letter].width = adjusted_width

# """
# Generate comparison results in XLSX format.
# Creates separate files for table and text differences.
# """

# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from pathlib import Path
# import json


# class ResultsGenerator:
#     """Generate XLSX output files with comparison results."""
    
#     def __init__(self, output_dir):
#         self.output_dir = Path(output_dir)
#         self.output_dir.mkdir(exist_ok=True)
    
#     def generate_table_results(self, differences, file_name_1, file_name_2):
#         """
#         Generate XLSX file with table differences.
        
#         Args:
#             differences: List of table differences
#             file_name_1: Name of first PDF
#             file_name_2: Name of second PDF
        
#         Returns:
#             Path to generated file
#         """
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Table Differences"
        
#         # Header styling
#         header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
#         header_font = Font(bold=True, color="FFFFFFFF")
#         thin_border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )
        
#         # Add metadata header
#         ws['A1'] = f"Comparison: {file_name_1} vs {file_name_2}"
#         ws['A1'].font = Font(bold=True, size=12)
#         ws.merge_cells('A1:D1')
        
#         # Column headers
#         headers = ["Source Location", "From File", "Comparison File", "Row Data"]
#         for col_idx, header in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col_idx)
#             cell.value = header
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = thin_border
        
#         # Add data rows
#         for row_idx, diff in enumerate(differences, 4):
#             source = ws.cell(row=row_idx, column=1)
#             source.value = diff["source"]
#             source.border = thin_border
#             source.alignment = Alignment(wrap_text=True)
            
#             from_file = ws.cell(row=row_idx, column=2)
#             from_file.value = diff["from_file"]
#             from_file.border = thin_border
            
#             comp_file = ws.cell(row=row_idx, column=3)
#             comp_file.value = diff["comparison_file"]
#             comp_file.border = thin_border
            
#             row_data = ws.cell(row=row_idx, column=4)
#             row_data.value = json.dumps(diff["row_data"], indent=2)
#             row_data.border = thin_border
#             row_data.alignment = Alignment(wrap_text=True, vertical="top")
        
#         # Adjust column widths
#         ws.column_dimensions['A'].width = 30
#         ws.column_dimensions['B'].width = 20
#         ws.column_dimensions['C'].width = 20
#         ws.column_dimensions['D'].width = 50
        
#         output_file = self.output_dir / "tables_differences.xlsx"
#         wb.save(output_file)
#         return output_file
    
#     def generate_text_results(self, differences, file_name_1, file_name_2):
#         """
#         Generate XLSX file with text differences.
        
#         Args:
#             differences: List of text differences
#             file_name_1: Name of first PDF
#             file_name_2: Name of second PDF
        
#         Returns:
#             Path to generated file
#         """
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Text Differences"
        
#         # Header styling
#         header_fill = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")
#         header_font = Font(bold=True, color="FFFFFFFF")
#         thin_border = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )
        
#         # Add metadata header
#         ws['A1'] = f"Comparison: {file_name_1} vs {file_name_2}"
#         ws['A1'].font = Font(bold=True, size=12)
#         ws.merge_cells('A1:D1')
        
#         # Column headers
#         headers = ["Source Location", "From File", "Comparison File", "Text Content"]
#         for col_idx, header in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col_idx)
#             cell.value = header
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#             cell.border = thin_border
        
#         # Add data rows
#         for row_idx, diff in enumerate(differences, 4):
#             source = ws.cell(row=row_idx, column=1)
#             source.value = diff["source"]
#             source.border = thin_border
#             source.alignment = Alignment(wrap_text=True)
            
#             from_file = ws.cell(row=row_idx, column=2)
#             from_file.value = diff["from_file"]
#             from_file.border = thin_border
            
#             comp_file = ws.cell(row=row_idx, column=3)
#             comp_file.value = diff["comparison_file"]
#             comp_file.border = thin_border
            
#             text_content = ws.cell(row=row_idx, column=4)
#             text_content.value = diff["text"]
#             text_content.border = thin_border
#             text_content.alignment = Alignment(wrap_text=True, vertical="top")
        
#         # Adjust column widths
#         ws.column_dimensions['A'].width = 30
#         ws.column_dimensions['B'].width = 20
#         ws.column_dimensions['C'].width = 20
#         ws.column_dimensions['D'].width = 60
        
#         output_file = self.output_dir / "text_differences.xlsx"
#         wb.save(output_file)
#         return output_file
    
#     def generate_comparison_summary(self, table_diff_count, text_diff_count, file_name_1, file_name_2):
#         """Generate summary log of comparison."""
#         summary = f"""
# COMPARISON SUMMARY
# ==================
# Date: [Generated]
# Comparing: {file_name_1} vs {file_name_2}

# Results:
# - Table Differences Found: {table_diff_count}
# - Text Differences Found: {text_diff_count}
# - Total Differences: {table_diff_count + text_diff_count}

# Output Files Generated:
# - tables_differences.xlsx (contains table rows unique to first PDF)
# - text_differences.xlsx (contains text paragraphs unique to first PDF)
# """
#         return summary

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

class OutputGenerator:
    """
    Generates clean, professional XLSX files with comparison results.
    Separates table and text differences clearly.
    """
    
    def __init__(self, output_dir, pdf1_name, pdf2_name):
        self.output_dir = output_dir
        self.pdf1_name = pdf1_name
        self.pdf2_name = pdf2_name
    
    def generate_table_results(self, different_rows):
        """
        Creates XLSX file for table differences.
        Shows only rows with NEW/DIFFERENT data.
        """
        if not different_rows:
            print("No table differences found.")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Table Differences"
        
        # Header styling
        header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add metadata header
        ws['A1'] = "PDF Comparison - Table Differences"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Source File: {self.pdf1_name}"
        ws['A3'] = f"Comparison File: {self.pdf2_name}"
        ws['A4'] = f"Only rows with NEW/DIFFERENT cells shown below"
        
        # Column headers
        headers = ["Table #", "Row #", "Cell 1", "Cell 2", "Cell 3", "Cell 4"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for row_num, row_data in enumerate(different_rows, 7):
            ws.cell(row=row_num, column=1).value = row_data['table_number']
            ws.cell(row=row_num, column=2).value = row_data['row_number']
            
            # Add cell data
            for col_num, cell_value in enumerate(row_data['row_data'], 3):
                ws.cell(row=row_num, column=col_num).value = str(cell_value)
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Save file
        output_file = f"{self.output_dir}/tables_comparison_{self.pdf1_name}_vs_{self.pdf2_name}.xlsx"
        wb.save(output_file)
        print(f"✓ Table comparison saved: {output_file}")
    
    def generate_text_results(self, different_sentences):
        """
        Creates XLSX file for text differences.
        Shows only sentences NOT found in comparison file.
        """
        if not different_sentences:
            print("No text differences found.")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Text Differences"
        
        # Header styling
        header_fill = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add metadata header
        ws['A1'] = "PDF Comparison - Text Differences"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')
        
        ws['A2'] = f"Source File: {self.pdf1_name}"
        ws['A3'] = f"Comparison File: {self.pdf2_name}"
        ws['A4'] = f"Only sentences NOT found in comparison file shown below"
        
        # Column headers
        headers = ["Sentence #", "Text Content", "Status"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for row_num, sentence_data in enumerate(different_sentences, 7):
            ws.cell(row=row_num, column=1).value = sentence_data['sentence_number']
            ws.cell(row=row_num, column=2).value = sentence_data['text']
            ws.cell(row=row_num, column=3).value = sentence_data['status']
            
            # Apply borders and wrapping
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).border = border
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 20
        
        # Save file
        output_file = f"{self.output_dir}/text_comparison_{self.pdf1_name}_vs_{self.pdf2_name}.xlsx"
        wb.save(output_file)
        print(f"✓ Text comparison saved: {output_file}")